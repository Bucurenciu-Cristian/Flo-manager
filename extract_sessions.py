# /// script
# dependencies = [
#   "openpyxl",
# ]
# ///

import openpyxl
import json
import copy
from datetime import datetime, date

# =============================================================================
# CONFIGURATION CONSTANTS - Modify these to control script behavior
# =============================================================================

# File paths
EXCEL_FILE_PATH = "excel.xlsx"
OUTPUT_FILE_LEGACY = "all_clients_sessions_final.json"  
OUTPUT_FILE_API = "fitness_sessions_api.json"

# Processing limits
MAX_CLIENTS = 200  # Maximum number of clients to process
START_FROM = 0     # Starting position (0-based index)
SEARCH_ROWS_PER_CLIENT = 80  # How many rows to search below each client name

# Date logic parameters
RECENT_THRESHOLD_DAYS = 90   # Days to consider "recent" for year determination
YEAR_THRESHOLD_DAYS = 183    # Days back to consider for previous year assignment

# Excel structure parameters
SESSION_COLUMNS_START = 4    # Column D (1-based: D=4)
SESSION_COLUMNS_END = 14     # Column M (1-based: M=13, range end is exclusive)
PREVIOUS_SESSIONS_SEARCH_START = 3  # Start searching 3 rows below client name
PREVIOUS_SESSIONS_SEARCH_END = 8    # End searching 7 rows below client name (exclusive)

# Previous sessions validation
MIN_PREVIOUS_SESSIONS = 30   # Minimum valid previous session count
MAX_PREVIOUS_SESSIONS = 1000 # Maximum valid previous session count

# Color codes (RGB values)
COLOR_PAID_GREEN = "FF00FF00"     # Green cells = paid sessions
COLOR_UNPAID_ORANGE = "FFFF9900"  # Orange cells = unpaid sessions
COLOR_DEFAULT_BLACK = "00000000"  # Default/empty color to skip
COLOR_DEFAULT_WHITE = "FFFFFFFF"  # Default/empty color to skip

# =============================================================================
# END CONFIGURATION
# =============================================================================

# NEW: global counter for processed clients
processed_count = 0  # Will be updated by extract_client_sessions

# Reference date is now dynamic (today)
CURRENT_DATE_REF = date.today()

# Helper to detect purely numeric strings (integers or floats like "230" or "230.0")
def _is_numeric_string(value: str) -> bool:
    txt = value.strip().replace(',', '').replace('.', '')
    return txt.isdigit()

def determine_year_from_date(day, month, current_date: date = None, threshold_days: int = YEAR_THRESHOLD_DAYS):
    """
    Determine the correct year for a session date given only day & month.

    Heuristic:
    1. Try the current year first. If that full date is **after** today → it must belong to the *previous* year.
    2. Otherwise, if that date is *far* in the past (older than `threshold_days`), assume it belongs to the previous year as well.
       This handles sessions logged more than ~6 months ago (e.g. today=18-Jun-2025, session 10-Jun could be 2024).
    3. In all other cases keep the current year.
    """
    if current_date is None:
        current_date = CURRENT_DATE_REF

    # Build candidate date in the current year
    try:
        candidate_this_year = date(current_date.year, month, day)
    except ValueError:
        # Fallback: if invalid (e.g. 31 Feb) just return current year
        return current_date.year

    # If candidate is in the *future* relative to reference date -> last year
    if candidate_this_year > current_date:
        return current_date.year - 1

    # If candidate is too far in the past -> last year
    if (current_date - candidate_this_year).days > threshold_days:
        return current_date.year - 1

    return current_date.year

def enhance_session_dates(sessions_list):
    """
    Enhance sessions with proper years and format as DD.MM.YYYY.

    Assumptions & rules:
    1. The raw list is ordered chronologically (oldest → newest) *within the sheet*.
    2. We start by assuming the first date belongs to *last* year.
    3. While walking forward, if a new (day,month) would go backwards in time,
       we bump its year forward until chronology is preserved.
    4. After the pass, if every date is still in that initial year **and** the
       timeline ends within `recent_threshold_days` of today, we realise the
       client actually has only current-year data → shift every date +1 year.
    """
    recent_threshold_days = RECENT_THRESHOLD_DAYS  # ~3 months
    enhanced = []

    if not sessions_list:
        return enhanced

    # Parse first date assuming previous year
    first_parts = sessions_list[0].split('.')
    if len(first_parts) != 2:
        return sessions_list  # fallback
    day1, month1 = map(int, first_parts)

    start_year = CURRENT_DATE_REF.year - 1
    dt_prev = datetime(start_year, month1, day1)
    enhanced.append(dt_prev)

    # Walk the rest
    for date_str in sessions_list[1:]:
        try:
            d, m = map(int, date_str.split('.'))
            year = dt_prev.year
            candidate = datetime(year, m, d)
            while candidate < dt_prev:
                year += 1
                candidate = datetime(year, m, d)
            dt_prev = candidate
            enhanced.append(candidate)
        except Exception:
            # Keep as-is if parse error
            enhanced.append(date_str)

    # Post-processing: if all datetimes, same year, and recent timeline → shift forward
    if all(isinstance(x, datetime) for x in enhanced):
        years = {x.year for x in enhanced}
        if len(years) == 1:
            year_only = next(iter(years))
            if (CURRENT_DATE_REF - enhanced[-1].date()).days <= recent_threshold_days:
                # Shift all forward by +1 year
                enhanced = [x.replace(year=x.year + 1) for x in enhanced]

    # Convert to strings, keep originals for any non-datetime entries
    result = []
    for item in enhanced:
        if isinstance(item, datetime):
            result.append(item.strftime("%d.%m.%Y"))
        else:
            result.append(item)
    return result

def find_previous_completed_sessions(ws, client_row, client_name):
    """Find optional previous completed sessions number in Column C, 3-7 rows below client name."""
    print(f"🔍 Searching for previous sessions in {client_name} section - Column C only, rows {client_row+3} to {client_row+7}")
    
    # Search only in Column C, 3-7 rows below client name
    candidates = []
    
    for check_row in range(client_row + PREVIOUS_SESSIONS_SEARCH_START, client_row + PREVIOUS_SESSIONS_SEARCH_END):
        col = 3  # Only Column C
        cell = ws.cell(row=check_row, column=col)
        if cell.value is not None:
            try:
                # Check if it's a standalone number (not a date)
                value_str = str(cell.value).strip()
                # More flexible number detection
                if value_str.replace('.', '').replace(',', '').isdigit():
                    num_val = int(float(value_str.replace(',', '.')))
                    if MIN_PREVIOUS_SESSIONS <= num_val <= MAX_PREVIOUS_SESSIONS:  # Broader reasonable range
                        candidates.append((num_val, check_row, col))
                        print(f"  Found candidate: {num_val} at row {check_row}, col {col}")
            except:
                continue
    
    # If we found candidates, prefer ones closer to the client row
    if candidates:
        # Sort by distance from client row
        candidates_sorted = sorted(candidates, key=lambda x: abs(x[1] - client_row))
        best_candidate = candidates_sorted[0]
        print(f"  Selected previous sessions: {best_candidate[0]} (Column C, row {best_candidate[1]})")
        return best_candidate[0]
    
    print(f"  No previous sessions number found for {client_name} in Column C")
    return 0

def extract_client_sessions(excel_file_path=EXCEL_FILE_PATH, max_clients=MAX_CLIENTS, start_from=START_FROM):
    """
    Extract client sessions from Excel file with proper color detection.
    
    Pattern:
    - Row X: "Numele" in column B, client name in column C
    - Rows below: colored cells in columns D-M (green=paid, orange=unpaid)
    - Dates can be below OR above colored cells
    - Green cells may contain text (stored in 'extra' property)
    """
    global processed_count  # Use the module-level counter so we can print it later
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active
    
    clients_data = {
        "clients": {},
        "updated": datetime.now().strftime("%Y-%m-%d"),
        "date_enhancement": {
            "enabled": True,
            "reference_date": "2025-06-18",
            "logic": "Dates after 18.6 are 2024, dates before/on 18.6 are 2025",
            "format": "DD.MM.YYYY"
        }
    }
    
    # Find all "Numele" positions (check both column B and C for client names)
    numele_positions = []
    processed_clients = set()  # Track processed clients to avoid duplicates
    
    for row in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row, column=2)
        cell_c = ws.cell(row=row, column=3)
        
        # Standard pattern: "Numele" in column B, client name in column C
        if cell_b.value and "Numele" in str(cell_b.value):
            client_name = cell_c.value
            if client_name and str(client_name).strip() and not _is_numeric_string(str(client_name)) and str(client_name).strip() not in processed_clients:
                numele_positions.append((row, str(client_name).strip()))
                processed_clients.add(str(client_name).strip())
        
        # Check column after "Varsta si Greutatea" for additional clients
        elif cell_b.value and "Varsta si Greutatea" in str(cell_b.value):
            # Check both column C and D for client name
            for check_col in [3, 4]:  # Check columns C and D
                cell_check = ws.cell(row=row, column=check_col)
                if cell_check.value and str(cell_check.value).strip() and not _is_numeric_string(str(cell_check.value)) and str(cell_check.value).strip() not in processed_clients:
                    client_name_candidate = str(cell_check.value).strip()
                    # Validate it's not a header or unwanted text
                    if not any(x in client_name_candidate.lower() for x in ['varsta', 'greutatea', 'motivatia', 'data']):
                        numele_positions.append((row, client_name_candidate))
                        processed_clients.add(client_name_candidate)
                        break  # Only take the first valid name found
        
        # Also check if there's a client name directly in column C without "Numele" marker
        elif cell_c.value and len(str(cell_c.value).strip()) > 2 and not _is_numeric_string(str(cell_c.value)):
            # This might be a client name in a different format
            potential_name = str(cell_c.value).strip()
            if potential_name not in processed_clients and not any(x in potential_name.lower() for x in ['numele', 'varsta', 'greutatea', 'data']):
                # Double check by looking for colored cells in nearby rows
                has_colored_cells = False
                for check_row in range(max(1, row-5), min(ws.max_row, row+10)):
                    for col in range(4, 14):
                        cell = ws.cell(row=check_row, column=col)
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                            rgb = str(cell.fill.fgColor.rgb)
                            if rgb in ["FF00FF00", "FFFF9900"] or "FF99" in rgb:
                                has_colored_cells = True
                                break
                    if has_colored_cells:
                        break
                
                if has_colored_cells:
                    numele_positions.append((row, potential_name))
                    processed_clients.add(potential_name)
    
    print(f"Found {len(numele_positions)} clients total, processing {min(max_clients, len(numele_positions)-start_from)} starting from position {start_from+1}")
    
    # Process clients from start_from to start_from + max_clients
    end_index = min(start_from + max_clients, len(numele_positions))
    processed_count = 0
    
    for i, (client_row, client_name) in enumerate(numele_positions[start_from:end_index]):
        processed_count += 1
        print(f"\n[{processed_count}/{min(max_clients, len(numele_positions)-start_from)}] Processing: {client_name} (row {client_row})")
        
        # Look for previous completed sessions in Column C, 3-7 rows below client name
        previous_completed = find_previous_completed_sessions(ws, client_row, client_name)
        print(f"Previous completed sessions: {previous_completed}")
        
        paid_sessions = []
        unpaid_sessions = []
        extra_data = []  # Store text from green cells
        undated_paid_count = 0  # Count green cells with no date (pre-paid sessions remaining)
        
        # Look for session data in the next rows after client name
        actual_index = start_from + i
        next_client_row = numele_positions[actual_index + 1][0] if actual_index + 1 < len(numele_positions) else ws.max_row
        search_end = min(client_row + SEARCH_ROWS_PER_CLIENT, next_client_row)
        print(f"  Search range: rows {client_row+1} to {search_end} (next client at {next_client_row})")
        
        # STEP 1: Find the first green cell with a date to establish reference point
        # Strategy: Find the first ROW that has any dates, then find the leftmost green cell in that row
        first_green_with_date_row = None
        first_green_with_date_col = None
        
        # Find the very FIRST date chronologically (earliest position)
        # Check both green (paid) and orange (unpaid) cells as valid starting points
        for row in range(client_row + 1, search_end):
            for col in range(SESSION_COLUMNS_START, SESSION_COLUMNS_END):  # Check columns D-M left to right
                # Check if there's a colored cell here
                cell = ws.cell(row=row, column=col)
                
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb)
                    
                    # Skip default/empty colors
                    if rgb == COLOR_DEFAULT_BLACK or rgb == COLOR_DEFAULT_WHITE:
                        continue
                    
                    # Check if it's a green cell (paid) OR orange cell (unpaid)
                    if rgb == COLOR_PAID_GREEN:
                        # Check if this green cell has a date
                        date_cell_below = ws.cell(row=row + 1, column=col)
                        date_cell_above = ws.cell(row=row - 1, column=col)
                        
                        if date_cell_below.value or date_cell_above.value:
                            first_green_with_date_row = row
                            first_green_with_date_col = col
                            date_value = date_cell_below.value or date_cell_above.value
                            print(f"  🎯 FIRST date found at row {row}, col {col}, date: {date_value}")
                            break
                    # Also check orange/yellow cells (unpaid) as potential starting points
                    elif rgb == COLOR_UNPAID_ORANGE or "FF99" in rgb or ("FFFF" in rgb[:4] and rgb != COLOR_DEFAULT_WHITE):
                        # Check if this orange cell has a date
                        date_cell_below = ws.cell(row=row + 1, column=col)
                        date_cell_above = ws.cell(row=row - 1, column=col)
                        
                        if date_cell_below.value or date_cell_above.value:
                            first_green_with_date_row = row
                            first_green_with_date_col = col
                            date_value = date_cell_below.value or date_cell_above.value
                            print(f"  🎯 FIRST date found at row {row}, col {col}, date: {date_value} (ORANGE)")
                            break
            
            if first_green_with_date_row:
                break
        
        if not first_green_with_date_row:
            print(f"  ❌ No green cells with dates found for {client_name}")
            # Create empty client data
            client_data = {
                "paid": [],
                "unpaid": [],
                "stats": {
                    "previous_completed": previous_completed,
                    "current_paid_used": 0,
                    "current_remaining": 0,
                    "current_unpaid": 0,
                    "total_current": 0,
                    "total_all_time": previous_completed,
                    # Legacy compatibility
                    "total": 0,
                    "paid": 0,
                    "paid_used": 0,
                    "paid_remaining": 0,
                    "unpaid": 0
                }
            }
            clients_data["clients"][client_name] = client_data
            print(f"  Summary: Previous={previous_completed}, Current=0 used + 0 remaining + 0 unpaid = 0, Total={previous_completed}")
            continue
        
        # STEP 2: Count sessions only from the first green cell with date onwards
        for row in range(client_row + 1, search_end):
            # Check for colored cells in columns D-M (4-13)
            colored_cells = []
            
            for col in range(SESSION_COLUMNS_START, SESSION_COLUMNS_END):  # Columns D-M
                cell = ws.cell(row=row, column=col)
                
                # Only count cells at or after the first green cell with date
                if row < first_green_with_date_row or (row == first_green_with_date_row and col < first_green_with_date_col):
                    continue
                
                # Simplified debug: just count
                # (removed detailed per-cell logging for cleaner output)
                
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb)
                    
                    # Skip default/empty colors
                    if rgb == COLOR_DEFAULT_BLACK or rgb == COLOR_DEFAULT_WHITE:
                        continue
                    
                    # Green cells = paid (FF00FF00)
                    if rgb == COLOR_PAID_GREEN:
                        cell_text_raw = cell.value
                        # Only keep non-numeric text. If the cell contains something that can be
                        # converted to a number (e.g. 30, 25.5, "30.0"), ignore it for the `extra` list.
                        if cell_text_raw is not None and str(cell_text_raw).strip():
                            cleaned_text = str(cell_text_raw).strip()
                            # Detect numeric values (integers or floats, dot or comma as decimal)
                            is_numeric = False
                            try:
                                # Replace comma with dot to support European decimals like "25,5"
                                float(cleaned_text.replace(",", "."))
                                # Succeeded => numeric
                                is_numeric = True
                            except ValueError:
                                # Not a pure number
                                pass
                            if not is_numeric:
                                colored_cells.append((col, "paid", cleaned_text))
                            else:
                                colored_cells.append((col, "paid", None))
                        else:
                            colored_cells.append((col, "paid", None))
                    # Orange/yellow cells = unpaid (FFFF9900 or similar)
                    elif rgb == COLOR_UNPAID_ORANGE or "FF99" in rgb or ("FFFF" in rgb[:4] and rgb != COLOR_DEFAULT_WHITE):
                        colored_cells.append((col, "unpaid", None))
            
            # If we found colored cells, check for dates (first below, then above if not found)
            if colored_cells:
                for col_data in colored_cells:
                    if len(col_data) == 3:
                        col, session_type, cell_text = col_data
                    else:
                        col, session_type = col_data[0], col_data[1]
                        cell_text = None
                    
                    # First check below (row + 1)
                    date_cell_below = ws.cell(row=row + 1, column=col)
                    date_cell_above = ws.cell(row=row - 1, column=col)
                    
                    date_found = False
                    
                    # Try below first
                    if date_cell_below.value:
                        date_cell = date_cell_below
                        date_found = True
                    # If not found below, try above
                    elif date_cell_above.value:
                        date_cell = date_cell_above
                        date_found = True
                    
                    if date_found:
                        date_str = str(date_cell.value).strip()
                        
                        # Convert date format from "2024-06-10 00:00:00" to "10.6"
                        try:
                            # Handle datetime objects or date strings
                            if hasattr(date_cell.value, 'date'):
                                # It's a datetime object
                                date_obj = date_cell.value.date()
                                day = date_obj.day
                                month = date_obj.month
                            elif "-" in date_str and len(date_str) >= 10:
                                # It's a string like "2024-06-10" or "2024-06-10 00:00:00"
                                date_part = date_str.split()[0]  # Remove time part if present
                                parts = date_part.split("-")
                                if len(parts) >= 3:
                                    day = int(parts[2])
                                    month = int(parts[1])
                                else:
                                    continue
                            else:
                                continue
                            
                            formatted_date = f"{day}.{month}"
                            
                            # Store session data
                            if session_type == "paid":
                                paid_sessions.append(formatted_date)
                                # Store extra text if present
                                if cell_text:
                                    extra_data.append({"date": formatted_date, "text": cell_text})
                            else:
                                unpaid_sessions.append(formatted_date)
                            
                            print(f"  Found {session_type} session: {formatted_date}" + (f" (extra: {cell_text})" if cell_text else ""))
                        except Exception as e:
                            print(f"  Could not parse date: {date_str} - {e}")
                    else:
                        # No date found – special handling for paid cells (green)
                        if session_type == "paid":
                            undated_paid_count += 1  # Treat as an already purchased session without specified date
                        print(f"  No date found for {session_type} cell in column {col}" + (" (counting as paid)" if session_type == "paid" else ""))
        
        # Enhance dates with proper years and chronological sorting
        enhanced_paid = enhance_session_dates(paid_sessions)
        enhanced_unpaid = enhance_session_dates(unpaid_sessions)
        
        # Calculate stats
        #   • previous_completed: sessions completed before this tracking method
        #   • paid_used: dated green cells (sessions already taken)
        #   • remaining: undated green cells (pre-paid sessions still available)
        #   • unpaid:    orange cells with dates (taken but unpaid)
        paid_used = len(enhanced_paid)
        remaining = undated_paid_count  # Exactly how many undated paid sessions are left
        total_paid_sessions = paid_used + remaining
        total_current = total_paid_sessions + len(enhanced_unpaid)
        total_all_time = previous_completed + total_current
        
        # Add client data with enhanced dates
        client_data = {
            "paid": enhanced_paid,
            "unpaid": enhanced_unpaid,
            "stats": {
                "previous_completed": previous_completed,
                "current_paid_used": paid_used,
                "current_remaining": remaining,
                "current_unpaid": len(enhanced_unpaid),
                "total_current": total_current,
                "total_all_time": total_all_time,
                # Legacy compatibility
                "total": total_current,
                "paid": total_paid_sessions,
                "paid_used": paid_used,
                "paid_remaining": remaining,
                "unpaid": len(enhanced_unpaid)
            }
        }
        
        # Add extra data if present
        if extra_data:
            client_data["extra"] = extra_data
        
        clients_data["clients"][client_name] = client_data
        
        extra_count = len(extra_data) if extra_data else 0
        print(f"  Summary: Previous={previous_completed}, Current={paid_used} used + {remaining} remaining + {len(enhanced_unpaid)} unpaid = {total_current}, Total={total_all_time}" + (f", {extra_count} with extra text" if extra_count > 0 else ""))
    
    return clients_data

def save_to_json(data, output_file="sessions_extracted.json"):
    """Save the extracted data to a JSON file optimized for Next.js frontend."""
    
    # Transform data structure for better frontend usability
    clients_array = []
    
    for client_name, client_data in data['clients'].items():
        # Create slug-style ID from name
        client_id = client_name.lower().replace(' ', '-').replace('ă', 'a').replace('â', 'a').replace('î', 'i').replace('ș', 's').replace('ț', 't')
        
        # Transform sessions to structured format
        paid_sessions = []
        unpaid_sessions = []
        
        # Handle both list and string formats
        if isinstance(client_data.get('paid'), list):
            for date_str in client_data['paid']:
                if date_str.strip():  # Skip empty dates
                    # Convert DD.MM.YYYY to YYYY-MM-DD for JavaScript Date compatibility
                    try:
                        parts = date_str.split('.')
                        if len(parts) == 3:
                            iso_date = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                            paid_sessions.append({
                                "date": iso_date,
                                "formatted": date_str
                            })
                    except:
                        # Fallback for malformed dates
                        paid_sessions.append({
                            "date": date_str,
                            "formatted": date_str
                        })
        elif isinstance(client_data.get('paid'), str) and client_data['paid'].strip():
            # Handle comma-separated string format
            for date_str in client_data['paid'].split(', '):
                if date_str.strip():
                    try:
                        parts = date_str.split('.')
                        if len(parts) == 3:
                            iso_date = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                            paid_sessions.append({
                                "date": iso_date,
                                "formatted": date_str
                            })
                    except:
                        paid_sessions.append({
                            "date": date_str,
                            "formatted": date_str
                        })
        
        # Same logic for unpaid sessions
        if isinstance(client_data.get('unpaid'), list):
            for date_str in client_data['unpaid']:
                if date_str.strip():
                    try:
                        parts = date_str.split('.')
                        if len(parts) == 3:
                            iso_date = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                            unpaid_sessions.append({
                                "date": iso_date,
                                "formatted": date_str
                            })
                    except:
                        unpaid_sessions.append({
                            "date": date_str,
                            "formatted": date_str
                        })
        elif isinstance(client_data.get('unpaid'), str) and client_data['unpaid'].strip():
            for date_str in client_data['unpaid'].split(', '):
                if date_str.strip():
                    try:
                        parts = date_str.split('.')
                        if len(parts) == 3:
                            iso_date = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                            unpaid_sessions.append({
                                "date": iso_date,
                                "formatted": date_str
                            })
                    except:
                        unpaid_sessions.append({
                            "date": date_str,
                            "formatted": date_str
                        })
        
        # Clean up stats - use camelCase and remove legacy fields
        stats = client_data.get('stats', {})
        clean_stats = {
            "previousCompleted": stats.get('previous_completed', 0),
            "currentPaidUsed": stats.get('current_paid_used', 0),
            "currentRemaining": stats.get('current_remaining', 0),
            "currentUnpaid": stats.get('current_unpaid', 0),
            "totalCurrent": stats.get('total_current', 0),
            "totalAllTime": stats.get('total_all_time', 0)
        }
        
        # Build client object
        client_obj = {
            "id": client_id,
            "name": client_name,
            "sessions": {
                "paid": paid_sessions,
                "unpaid": unpaid_sessions
            },
            "stats": clean_stats,
            "lastUpdated": data.get('updated', datetime.now().strftime("%Y-%m-%d"))
        }
        
        # Add extra data if present
        if 'extra' in client_data:
            client_obj['extra'] = client_data['extra']
        
        clients_array.append(client_obj)
    
    # Create frontend-optimized structure
    frontend_data = {
        "clients": clients_array,
        "metadata": {
            "totalClients": len(clients_array),
            "generatedAt": datetime.now().isoformat() + "Z",
            "version": "1.0",
            "dateEnhancement": data.get('date_enhancement', {})
        }
    }

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(frontend_data, f, indent=2, ensure_ascii=False)
    print(f"\nData saved to {output_file}")

if __name__ == "__main__":
    try:
        print("Extracting session data for ALL clients...")
        session_data = extract_client_sessions()
        
        # Save to JSON with enhanced dates  
        save_to_json(session_data, OUTPUT_FILE_API)
        
        # Print enhancement summary
        total_clients = len(session_data['clients'])
        total_sessions = sum(info['stats']['total'] for info in session_data['clients'].values())
        year_2024_count = 0
        year_2025_count = 0
        
        for client_data in session_data['clients'].values():
            for session in client_data.get('paid', []) + client_data.get('unpaid', []):
                if '2024' in session:
                    year_2024_count += 1
                elif '2025' in session:
                    year_2025_count += 1
        
        print("\n🎯 EXTRACTION & ENHANCEMENT SUMMARY")
        print("=" * 50)
        print(f"✅ Total clients processed: {total_clients}")
        print(f"✅ Total sessions extracted: {total_sessions}")
        if total_sessions > 0:
            print(f"📅 Sessions from 2024: {year_2024_count} ({year_2024_count/total_sessions*100:.1f}%)")
            print(f"📅 Sessions from 2025: {year_2025_count} ({year_2025_count/total_sessions*100:.1f}%)")
        else:
            print(f"📅 Sessions from 2024: {year_2024_count} (0.0%)")
            print(f"📅 Sessions from 2025: {year_2025_count} (0.0%)")
        print(f"🗓️  Enhanced format: DD.MM.YYYY")
        print(f"📊 Logic: Dates after 18.6 → 2024, dates before/on 18.6 → 2025")
        
        print(f"\n📋 Sample clients:")
        sample_count = 0
        clients_with_extra = 0
        for name, info in session_data['clients'].items():
            if 'extra' in info:
                clients_with_extra += 1
            if sample_count >= 3:
                continue
            stats = info['stats']
            if stats['total'] > 0:
                paid_sample = info['paid'][:2] if info['paid'] else []
                unpaid_sample = info['unpaid'][:1] if info['unpaid'] else []
                extra_sample = info.get('extra', [])[:1] if info.get('extra') else []
                print(f"  {name}: {stats['total']} sessions")
                if paid_sample:
                    print(f"    Paid: {paid_sample}")
                if unpaid_sample:
                    print(f"    Unpaid: {unpaid_sample}")
                if extra_sample:
                    print(f"    Extra: {extra_sample}")
                sample_count += 1
        
        print(f"\n📝 Clients with extra text: {clients_with_extra}")
        print(f"🔢 Total clients processed: {processed_count}")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()