# /// script
# dependencies = [
#   "pandas",
#   "openpyxl",
# ]
# ///

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import json
from datetime import datetime
import re

def extract_training_data(excel_file_path):
    """
    Extract training session data from Excel file and convert to JSON format.
    
    Args:
        excel_file_path (str): Path to the Excel file
    
    Returns:
        dict: JSON structure with all client training data
    """
    
    # Load the workbook with openpyxl to access cell colors
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active
    
    # Current date for year determination
    current_date = datetime(2025, 6, 18)
    
    clients_data = {
        "trainer": {
            "businessName": "Prezenta GRUP",
            "year": 2025
        },
        "clients": [],
        "sessionStatuses": {
            "paid": {
                "description": "Session completed and paid for",
                "color": "green"
            },
            "unpaid": {
                "description": "Session completed but not yet paid",
                "color": "orange"
            }
        },
        "metadata": {
            "lastUpdated": current_date.strftime("%Y-%m-%d"),
            "currency": "RON",
            "packageSize": 10,
            "pricePerSession": 30
        }
    }
    
    # Find client sections (assuming each client has 3 rows: Numele, Varsta si Greutatea, Motivatia)
    client_id = 1
    row = 1
    
    while row <= ws.max_row:
        # Look for "Numele" cell which indicates start of a client section
        cell_a = ws.cell(row=row, column=1)
        
        if cell_a.value and "Numele" in str(cell_a.value):
            # Get client name from column B
            client_name = ws.cell(row=row, column=2).value
            
            if client_name:
                print(f"Processing client: {client_name}")
                
                client_data = {
                    "id": f"client_{client_id:03d}",
                    "name": client_name,
                    "sessions": []
                }
                
                # Process session rows (starting from row+3, skipping the header rows)
                session_row = row + 3
                
                # Continue until we hit an empty row or another client section
                while session_row <= ws.max_row:
                    # Check if this row has session data
                    has_data = False
                    
                    # Check columns C through T (3 through 20) for session dates
                    for col in range(3, 21):  # Adjust range based on your Excel
                        cell = ws.cell(row=session_row, column=col)
                        
                        if cell.value:
                            has_data = True
                            
                            # Determine if session is paid based on cell color
                            fill = cell.fill
                            is_paid = False
                            
                            if fill and fill.patternType:
                                # Check for green color (paid)
                                if fill.fgColor and fill.fgColor.rgb:
                                    # Green colors typically have high green component
                                    rgb = fill.fgColor.rgb
                                    if len(str(rgb)) >= 6:
                                        # Extract RGB values
                                        try:
                                            if isinstance(rgb, str):
                                                r = int(rgb[2:4], 16) if len(rgb) >= 4 else 0
                                                g = int(rgb[4:6], 16) if len(rgb) >= 6 else 0
                                                b = int(rgb[6:8], 16) if len(rgb) >= 8 else 0
                                            else:
                                                # Handle ARGB format
                                                rgb_str = str(rgb)
                                                r = int(rgb_str[-6:-4], 16)
                                                g = int(rgb_str[-4:-2], 16)
                                                b = int(rgb_str[-2:], 16)
                                            
                                            # Green detection: high green, low red and blue
                                            if g > 200 and r < 100 and b < 100:
                                                is_paid = True
                                        except:
                                            # If color detection fails, assume unpaid
                                            is_paid = False
                            
                            # Parse the date value
                            date_str = str(cell.value)
                            
                            # Handle special date formats
                            if "2024" in date_str:
                                # Already has year
                                parts = date_str.replace(".2024", "").split(".")
                                if len(parts) == 2:
                                    day, month = int(parts[0]), int(parts[1])
                                    year = 2024
                            else:
                                # Parse day.month format
                                parts = date_str.split(".")
                                if len(parts) == 2:
                                    day, month = int(parts[0]), int(parts[1])
                                    
                                    # Determine year based on current date
                                    if month < 6 or (month == 6 and day <= 18):
                                        year = 2025
                                    else:
                                        year = 2024
                            
                            # Create session entry
                            session = {
                                "date": f"{year}-{month:02d}-{day:02d}",
                                "status": "paid" if is_paid else "unpaid",
                                "displayDate": date_str
                            }
                            
                            client_data["sessions"].append(session)
                    
                    if not has_data:
                        break
                    
                    session_row += 1
                
                # Sort sessions by date
                client_data["sessions"].sort(key=lambda x: x["date"])
                
                # Calculate summary
                total_sessions = len(client_data["sessions"])
                paid_sessions = sum(1 for s in client_data["sessions"] if s["status"] == "paid")
                unpaid_sessions = total_sessions - paid_sessions
                
                # Calculate remaining sessions based on packages of 10
                sessions_used = total_sessions
                packages_used = (sessions_used + 9) // 10  # Round up
                total_purchased = packages_used * 10
                remaining_sessions = total_purchased - sessions_used
                
                client_data["summary"] = {
                    "totalSessions": total_sessions,
                    "paidSessions": paid_sessions,
                    "unpaidSessions": unpaid_sessions,
                    "remainingSessions": remaining_sessions,
                    "packagesUsed": packages_used
                }
                
                clients_data["clients"].append(client_data)
                client_id += 1
                
                # Skip to next potential client (current row + 4 at minimum)
                row += 4
            else:
                row += 1
        else:
            row += 1
    
    return clients_data

def save_to_json(data, output_file="training_sessions.json"):
    """Save the extracted data to a JSON file."""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Data saved to {output_file}")

def print_summary(data):
    """Print a summary of the extracted data."""
    print("\n=== EXTRACTION SUMMARY ===")
    print(f"Total clients: {len(data['clients'])}")
    
    for client in data['clients']:
        print(f"\n{client['name']}:")
        print(f"  - Total sessions: {client['summary']['totalSessions']}")
        print(f"  - Paid sessions: {client['summary']['paidSessions']}")
        print(f"  - Unpaid sessions: {client['summary']['unpaidSessions']}")
        print(f"  - Remaining sessions: {client['summary']['remainingSessions']}")
        print(f"  - Packages used: {client['summary']['packagesUsed']}")

# Example usage
if __name__ == "__main__":
    # Using the exact filename
    excel_file = "excel.xlsx"
    
    try:
        # Extract data
        training_data = extract_training_data(excel_file)
        
        # Save to JSON
        save_to_json(training_data)
        
        # Print summary
        print_summary(training_data)
        
    except FileNotFoundError:
        print(f"Error: Could not find file '{excel_file}'")
        print("Please make sure the file 'excel.xlsx' exists in the current directory.")
    except Exception as e:
        print(f"Error processing file: {e}")
        import traceback
        traceback.print_exc()